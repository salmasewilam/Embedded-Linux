#Python program to parse header file and read all prototypes of function 
#and insert it into excel sheet with unique ID start with IDX0
import openpyxl

workbook=openpyxl.Workbook()
new_sheet= workbook.create_sheet(title="newsheet.xlsx")

file=open("headerfile.txt.txt","r")

startvalue="IDX00"
i=1
for x in file:
    new_sheet.cell(row=i, column=1).value = x
    new_sheet.cell(row=i, column=2).value = startvalue+i
    i=i+1
	
workbook.save("newsheet.xlsx")